import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl  # For more advanced Excel manipulation

def scrape_cafe_menu(url):
    """
    Scrapes the menu from the given Monkey Tribe Cafe Vagamon URL.

    Args:
        url (str): The URL of the cafe's webpage.

    Returns:
        dict: A dictionary where keys are menu category headings and
              values are lists of dictionaries, each containing 'item' and 'price'.
              Returns None if the request fails.
    """
    try:
        response = requests.get(url)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        menu_data = {}

        # Find all spans with the class wixui-rich-text__text
        text_spans = soup.find_all('span', class_='wixui-rich-text__text')

        # Iterate through the spans
        category = None
        for span in text_spans:
            # Check for category (appetizers, main course etc.)
            if 'font-family:trend-sans-w00-four' in span.get('style', ''):
                category = span.text.strip()
                menu_data[category] = []
            # Check for item name or price
            elif 'font-family:cinzel' in span.get('style', ''):
                if category:
                    text = span.text.strip()
                    if '₹' in text:
                        price = text
                        if (len(menu_data[category]) > 0):
                           menu_data[category][-1]['price'] = price
                    else:
                        item_name = text
                        menu_data[category].append({'item': item_name, 'price': ''})

        return menu_data

    except requests.exceptions.RequestException as e:
        print(f"Error fetching the webpage: {e}")
        return None



def save_menu_to_excel(menu_data, filename="menu.xlsx"):
    """
    Saves the menu data to an Excel file with improved formatting.

    Args:
        menu_data (dict): A dictionary containing the menu data.
        filename (str, optional): The name of the Excel file to save to.
            Defaults to "menu.xlsx".
    """
    try:
        # Create an empty list to hold the data for the DataFrame
        excel_data = []

        # Iterate through the menu data to structure it for Excel
        for category, items in menu_data.items():
            # Add the category as a row, with a special identifier.
            excel_data.append({'Name': category, 'Price': None, 'Category': 'Category'})
            for item in items:
                excel_data.append({'Name': item['item'], 'Price': item['price'], 'Category': 'Item'})

        # Create a Pandas DataFrame
        df = pd.DataFrame(excel_data)

        # Save the DataFrame to an Excel file
        df.to_excel(filename, index=False)

        # --- Formatting with openpyxl ---
        # Load the workbook and worksheet
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active

        # Iterate through rows and format category rows
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            if row[2].value == 'Category':  # If the 'Category' column is 'Category'
                # Make the entire row bold
                for cell in row:
                    cell.font = openpyxl.styles.Font(bold=True)
                # Align category to the left
                row[0].alignment = openpyxl.styles.Alignment(horizontal='left')

        # Save the formatted workbook
        workbook.save(filename)

        print(f"Menu data successfully saved and formatted to {filename}")

    except Exception as e:
        print(f"An error occurred: {e}")
        return None



if __name__ == "__main__":
    cafe_url = 'https://www.monkeytribe.com/cafevagamon'
    menu = scrape_cafe_menu(cafe_url)

    if menu:
        save_menu_to_excel(menu)
    else:
        print("Failed to retrieve the menu.")
